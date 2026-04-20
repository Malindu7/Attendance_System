from flask import Flask, render_template, Response, request, redirect, url_for, session, flash, send_file
import cv2
import face_recognition
import numpy as np
import os
from datetime import datetime
from pymongo import MongoClient
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)
app.secret_key = "kdu_group04_secret"

# ============================================================
# CONFIGURATION — change IMAGE_FOLDER to your path
# ============================================================
IMAGE_FOLDER   =  "D:\\Lecture Notes\\4th Year\\7th semester\\IPCV\\group project\\preprocessing\\images"
ADMIN_EMAIL    = "admin@kdu.ac.lk"
ADMIN_PASSWORD = "123"

# ============================================================
# COURSE & SUBJECT STRUCTURE
# ============================================================
COURSES = {
    "Data Science": [
        "Digital Image Processing",
        "Machine Learning",
        "Data Mining & Analytics",
    ],
    "Computer Science": [
        "Computer Vision",
        "Algorithm Design",
        "Operating Systems",
    ],
    "Computer Engineering": [
        "Embedded Systems",
        "Digital Signal Processing",
        "Computer Architecture",
    ],
    "Software Engineering": [
        "Software Architecture",
        "Agile Development",
        "Database Systems",
    ],
}

# ============================================================
# MONGODB
# ============================================================
client = MongoClient("mongodb://localhost:27017/")
db     = client['attendance_db']

# Two collections:
#   students  — stores registered student info + enrolled course
#   logs      — one attendance record per student per subject per day
students_col   = db['students']
attendance_col = db['logs']

# ============================================================
# GLOBAL FACE DATA
# ============================================================
camera               = None
known_face_encodings = []
known_face_names     = []   # stored as  "Name_StudentID"
known_face_courses   = []   # enrolled course for each face


# ============================================================
# PREPROCESSING HELPERS
# ============================================================
def apply_clahe(frame_rgb):
    lab = cv2.cvtColor(frame_rgb, cv2.COLOR_RGB2LAB)
    l, a, b = cv2.split(lab)
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
    l_eq  = clahe.apply(l)
    return cv2.cvtColor(cv2.merge([l_eq, a, b]), cv2.COLOR_LAB2RGB)

def apply_gaussian_blur(frame_rgb):
    return cv2.GaussianBlur(frame_rgb, (3, 3), sigmaX=0.5)

def preprocess_frame(frame_rgb):
    """CLAHE → Gaussian Blur → resize to 25%"""
    return cv2.resize(apply_gaussian_blur(apply_clahe(frame_rgb)), (0, 0), fx=0.25, fy=0.25)


# ============================================================
# LOAD KNOWN FACES FROM DISK
# Image filenames:  Name_StudentID_Course.jpg
#   e.g.  John_Doe_D-BCS-24-0010_Computer Science.jpg
# ============================================================
def load_known_faces():
    global known_face_encodings, known_face_names, known_face_courses
    known_face_encodings = []
    known_face_names     = []
    known_face_courses   = []

    if not os.path.exists(IMAGE_FOLDER):
        os.makedirs(IMAGE_FOLDER)

    for filename in os.listdir(IMAGE_FOLDER):
        if not filename.lower().endswith(('.jpg', '.jpeg', '.png')):
            continue
        filepath = os.path.join(IMAGE_FOLDER, filename)
        img = cv2.imread(filepath)
        if img is None:
            continue

        img_rgb   = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        img_ready = apply_gaussian_blur(apply_clahe(img_rgb))

        try:
            enc = face_recognition.face_encodings(img_ready)[0]
        except IndexError:
            print(f"[WARN] No face found in {filename}")
            continue

        # Parse filename: Name--StudentID--Course.jpg  (new format)
        # Also handles legacy Name_StudentID_Course.jpg
        stem = os.path.splitext(filename)[0]

        if '--' in stem:
            parts  = stem.split('--')
            label  = parts[0] if len(parts) >= 1 else stem
            course = parts[2].replace('-', ' ') if len(parts) >= 3 else "Unknown"
        else:
            # Legacy format — split from the RIGHT so names with _ are safe
            parts = stem.rsplit('_', 2)
            if len(parts) == 3:
                label  = parts[0]
                course = parts[2].replace('-', ' ')
            elif len(parts) == 2:
                label  = parts[0]
                course = parts[1].replace('-', ' ')
            else:
                label  = stem
                course = "Unknown"

        known_face_encodings.append(enc)
        known_face_names.append(label)
        known_face_courses.append(course)
        print(f"[OK] {label}  course={course}")

    print(f"[INFO] {len(known_face_names)} faces loaded.")

load_known_faces()


# ============================================================
# CAMERA HELPERS
# ============================================================
def open_camera():
    global camera
    if camera is not None and camera.isOpened():
        camera.release()
    import time; time.sleep(0.5)
    camera = cv2.VideoCapture(0, cv2.CAP_DSHOW)
    if not camera.isOpened():
        time.sleep(1)
        camera = cv2.VideoCapture(0)
    return camera


# ============================================================
# FRAME GENERATOR — VERIFY MODE
# Marks attendance only if:
#   1. Student is enrolled in the correct course for this subject
#   2. No record exists for this student+subject today
# ============================================================
def gen_frames_verify(subject, course):
    cam = open_camera()
    while True:
        ok, frame = cam.read()
        if not ok:
            break

        frame_rgb  = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        small      = preprocess_frame(frame_rgb)
        face_locs  = face_recognition.face_locations(small, model='hog')
        face_encs  = face_recognition.face_encodings(small, face_locs)

        for enc, loc in zip(face_encs, face_locs):
            name  = "Unknown"
            color = (0, 0, 220)
            note  = ""

            dists = face_recognition.face_distance(known_face_encodings, enc)
            if len(dists) > 0:
                idx  = int(np.argmin(dists))
                dist = dists[idx]
                if dist < 0.50:
                    candidate_name   = known_face_names[idx]
                    candidate_course = known_face_courses[idx]

                    if candidate_course.lower() != course.lower():
                        # Wrong course — show red + message
                        name  = candidate_name.upper()
                        color = (0, 60, 200)
                        note  = f"Not enrolled: {candidate_course}"
                    else:
                        name  = candidate_name.upper()
                        color = (0, 200, 0)

                        # One record per student per subject per day
                        today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
                        exists = attendance_col.find_one({
                            "name":    name,
                            "subject": subject,
                            "timestamp": {"$gte": today}
                        })
                        if exists is None:
                            attendance_col.insert_one({
                                "name":      name,
                                "subject":   subject,
                                "course":    course,
                                "timestamp": datetime.now()
                            })
                            print(f"[LOG] {name} | {subject} | {course}")
                        else:
                            note = "Already marked today"

            top, right, bottom, left = [v * 4 for v in loc]
            cv2.rectangle(frame, (left, top), (right, bottom), color, 2)
            cv2.rectangle(frame, (left, bottom - 35), (right, bottom), color, cv2.FILLED)
            cv2.putText(frame, name, (left + 6, bottom - 6),
                        cv2.FONT_HERSHEY_DUPLEX, 0.65, (255, 255, 255), 1)
            if note:
                cv2.putText(frame, note, (left, top - 8),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.48, (0, 220, 255), 1)

        cv2.putText(frame, f"{course} | {subject}", (10, 28),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 230, 0), 2)
        cv2.putText(frame, f"Faces registered: {len(known_face_names)}", (10, 56),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.5, (200, 200, 0), 1)

        _, buf = cv2.imencode('.jpg', frame)
        yield (b'--frame\r\nContent-Type: image/jpeg\r\n\r\n' + buf.tobytes() + b'\r\n')


# ============================================================
# FRAME GENERATOR — REGISTER MODE
# ============================================================
def gen_frames_register():
    cam = open_camera()
    while True:
        ok, frame = cam.read()
        if not ok:
            break
        cv2.putText(frame, "Position face & click Capture", (10, 30),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 165, 255), 2)
        _, buf = cv2.imencode('.jpg', frame)
        yield (b'--frame\r\nContent-Type: image/jpeg\r\n\r\n' + buf.tobytes() + b'\r\n')


# ============================================================
# PREPROCESSING DEMO — ALL 11 STAGES
# ============================================================
def gen_preprocessing_demo():
    global camera
    if camera is not None and camera.isOpened():
        camera.release()
    import time; time.sleep(0.5)
    cap = cv2.VideoCapture(0, cv2.CAP_DSHOW)
    if not cap.isOpened():
        time.sleep(1)
        cap = cv2.VideoCapture(0)

    TW, TH = 320, 240

    def label(img, text, col=(255, 255, 255)):
        out = img.copy()
        cv2.rectangle(out, (0, 0), (TW, 32), (0, 0, 0), cv2.FILLED)
        cv2.putText(out, text, (6, 22), cv2.FONT_HERSHEY_SIMPLEX, 0.52, col, 1)
        return out

    def hdiv(w):
        d = np.zeros((4, w, 3), dtype=np.uint8); d[:] = (20, 80, 180); return d

    def hdr_bar(w):
        t = np.zeros((36, w, 3), dtype=np.uint8); t[:] = (8, 12, 35)
        cv2.putText(t, "11-STAGE PREPROCESSING PIPELINE",
                    (8, 24), cv2.FONT_HERSHEY_SIMPLEX, 0.48, (80, 180, 255), 1)
        return t

    def enc_chart(encoding):
        chart = np.zeros((TH, TW, 3), dtype=np.uint8); chart[:] = (8, 8, 28)
        if encoding is not None:
            v = np.array(encoding)
            mn, mx = v.min(), v.max()
            norm = (v - mn) / (mx - mn + 1e-9)
            bw = max(1, TW // 128)
            for i, val in enumerate(norm):
                x  = i * bw
                bh = int(val * (TH - 52))
                g  = int(val * 200)
                b  = int((1 - val) * 200)
                cv2.rectangle(chart, (x, TH - 52 - bh), (x + bw - 1, TH - 52), (b, g + 55, 255 - b), -1)
            cv2.putText(chart, "128-D Encoding Vector", (4, TH - 34),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.42, (150, 150, 255), 1)
            cv2.putText(chart, f"min:{mn:.2f}  max:{mx:.2f}", (4, TH - 14),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.38, (100, 100, 180), 1)
        else:
            cv2.putText(chart, "No face detected", (TW // 2 - 70, TH // 2),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.5, (60, 60, 100), 1)
        return chart

    def match_panel(matched, dist, found):
        p = np.zeros((TH, TW, 3), dtype=np.uint8); p[:] = (8, 8, 28)
        if not found:
            cv2.putText(p, "No face in frame", (20, TH // 2),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.5, (60, 60, 100), 1)
            return p
        if matched and dist is not None and dist < 0.50:
            col, verdict = (0, 210, 110), "VERIFIED"
        else:
            col, verdict = (0, 70, 210), "UNKNOWN"
        cv2.rectangle(p, (10, 20), (TW - 10, 78), col, 2)
        tw = cv2.getTextSize(verdict, cv2.FONT_HERSHEY_DUPLEX, 0.9, 2)[0][0]
        cv2.putText(p, verdict, ((TW - tw) // 2, 62), cv2.FONT_HERSHEY_DUPLEX, 0.9, col, 2)
        label_str = (matched or "Not registered")[:22]
        cv2.putText(p, label_str, (10, 108), cv2.FONT_HERSHEY_SIMPLEX, 0.44, (200, 200, 255), 1)
        if dist is not None:
            cv2.putText(p, f"Distance: {dist:.4f}   Threshold: 0.50",
                        (10, 136), cv2.FONT_HERSHEY_SIMPLEX, 0.38, col, 1)
            total = TW - 20
            filled = int(min(dist / 0.8, 1.0) * total)
            cv2.rectangle(p, (10, 148), (10 + total, 162), (35, 35, 70), -1)
            cv2.rectangle(p, (10, 148), (10 + filled, 162), col, -1)
            tx = int(0.50 / 0.8 * total) + 10
            cv2.line(p, (tx, 143), (tx, 166), (0, 229, 255), 2)
            cv2.putText(p, "0.50", (tx - 12, 178),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.36, (0, 229, 255), 1)
        return p

    while True:
        ok, frame_bgr = cap.read()
        if not ok:
            break
        h, w = frame_bgr.shape[:2]

        # 1 Raw BGR
        s1 = label(cv2.resize(frame_bgr, (TW, TH)), "1. Raw BGR Frame")
        # 2 BGR→RGB
        frame_rgb = cv2.cvtColor(frame_bgr, cv2.COLOR_BGR2RGB)
        s2 = label(cv2.cvtColor(cv2.resize(frame_rgb, (TW, TH)), cv2.COLOR_RGB2BGR), "2. BGR to RGB")
        # 3 CLAHE
        frame_clahe = apply_clahe(frame_rgb)
        s3 = label(cv2.cvtColor(cv2.resize(frame_clahe, (TW, TH)), cv2.COLOR_RGB2BGR), "3. CLAHE", (0, 220, 255))
        # 4 Gaussian blur
        frame_blur = apply_gaussian_blur(frame_clahe)
        s4 = label(cv2.cvtColor(cv2.resize(frame_blur, (TW, TH)), cv2.COLOR_RGB2BGR), "4. Gaussian Blur")
        # 5 Grayscale
        gray = cv2.cvtColor(frame_blur, cv2.COLOR_RGB2GRAY)
        s5 = label(cv2.cvtColor(cv2.resize(gray, (TW, TH)), cv2.COLOR_GRAY2BGR), "5. Grayscale")
        # 6 HOG detection
        small = cv2.resize(frame_blur, (0, 0), fx=0.25, fy=0.25)
        face_locs = face_recognition.face_locations(small, model='hog')
        s6_base = cv2.resize(frame_bgr, (TW, TH))
        sx, sy = TW / w, TH / h
        for (t, r, b2, l) in face_locs:
            cv2.rectangle(s6_base,
                          (int(l*4*sx), int(t*4*sy)), (int(r*4*sx), int(b2*4*sy)), (0, 255, 0), 2)
        s6 = label(s6_base, "6. HOG Detection", (0, 255, 0))

        # Blank panels as defaults for stages 7-9
        blank = np.zeros((TH, TW, 3), dtype=np.uint8); blank[:] = (8, 8, 28)
        s7, s8, s9 = [label(blank.copy(), f"{n}. No face") for n in ["7", "8", "9"]]
        encoding = None; matched_name = None; best_dist = None; face_found = len(face_locs) > 0

        if face_locs:
            top, right, bottom, lft = face_locs[0]
            pad = 20
            ft = max(0, top*4 - pad);  fl = max(0, lft*4 - pad)
            fb = min(h, bottom*4 + pad); fr = min(w, right*4 + pad)

            # 7 ROI crop
            roi_bgr = frame_bgr[ft:fb, fl:fr]
            if roi_bgr.size > 0:
                roi_r = cv2.resize(roi_bgr, (TW, TH))
                cv2.rectangle(roi_r, (0, 0), (TW, TH), (0, 180, 255), 3)
                cv2.putText(roi_r, f"{fr-fl}x{fb-ft}px +20pad", (6, TH-10),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.38, (0, 180, 255), 1)
                s7 = label(roi_r, "7. ROI Crop", (0, 180, 255))

            # 8 Face alignment landmarks
            roi_rgb_crop = frame_rgb[ft:fb, fl:fr]
            if roi_rgb_crop.size > 0:
                s8_img = cv2.cvtColor(roi_rgb_crop, cv2.COLOR_RGB2BGR)
                s8_img = cv2.resize(s8_img, (TW, TH))
                rx, ry = TW / (fr - fl), TH / (fb - ft)
                lm_list = face_recognition.face_landmarks(
                    frame_rgb, [(top*4, right*4, bottom*4, lft*4)])
                if lm_list:
                    lm = lm_list[0]
                    lm_col = {'left_eye':(255,80,80),'right_eye':(80,80,255),
                              'nose_tip':(80,255,80),'top_lip':(255,160,0),
                              'bottom_lip':(255,160,0),'left_eyebrow':(180,0,180),
                              'right_eyebrow':(180,0,180)}
                    for feat, pts in lm.items():
                        c = lm_col.get(feat, (180, 180, 180))
                        for px, py in pts:
                            cv2.circle(s8_img, (int((px-fl)*rx), int((py-ft)*ry)), 3, c, -1)
                    le = np.mean(lm.get('left_eye',  [(0, 0)]), axis=0)
                    re = np.mean(lm.get('right_eye', [(10, 0)]), axis=0)
                    lx, ly   = int((le[0]-fl)*rx), int((le[1]-ft)*ry)
                    rx2, ry2 = int((re[0]-fl)*rx), int((re[1]-ft)*ry)
                    cv2.line(s8_img, (lx, ly), (rx2, ry2), (0, 255, 255), 2)
                    angle = np.degrees(np.arctan2(re[1]-le[1], re[0]-le[0]))
                    cv2.putText(s8_img, f"angle:{angle:.1f}deg", (4, TH-10),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.38, (0, 255, 255), 1)
                s8 = label(s8_img, "8. Face Alignment", (255, 255, 0))

            # 9 Normalisation heatmap
            if roi_rgb_crop.size > 0:
                n = roi_rgb_crop.astype(np.float32) / 255.0
                n = (n - np.array([0.485, 0.456, 0.406])) / np.array([0.229, 0.224, 0.225])
                vis = ((n - n.min()) / (n.max() - n.min() + 1e-9) * 255).astype(np.uint8)
                vis = cv2.applyColorMap(cv2.cvtColor(
                    cv2.cvtColor(vis, cv2.COLOR_RGB2BGR), cv2.COLOR_BGR2GRAY), cv2.COLORMAP_PLASMA)
                vis = cv2.resize(vis, (TW, TH))
                cv2.putText(vis, "÷255, ImageNet mean/std", (4, TH-10),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.38, (255, 200, 100), 1)
                s9 = label(vis, "9. Pixel Normalisation", (255, 180, 0))

            # 10-11 encoding & match
            try:
                encs = face_recognition.face_encodings(small, face_locs)
                if encs:
                    encoding = encs[0]
                    if known_face_encodings:
                        dists    = face_recognition.face_distance(known_face_encodings, encoding)
                        best_idx = int(np.argmin(dists))
                        best_dist = float(dists[best_idx])
                        if best_dist < 0.50:
                            matched_name = known_face_names[best_idx]
            except Exception:
                pass

        s10 = label(enc_chart(encoding), "10. ResNet-34  128-D Vector", (180, 100, 255))
        s11 = label(match_panel(matched_name, best_dist, face_found),
                    "11. Euclidean Match",
                    (0, 220, 120) if (matched_name and best_dist and best_dist < 0.50) else (80, 120, 255))

        W3   = TW * 3
        pad_p = np.zeros((TH, TW, 3), dtype=np.uint8); pad_p[:] = (6, 6, 18)
        demo  = np.vstack([
            hdr_bar(W3),
            np.hstack([s1, s2, s3]),  hdiv(W3),
            np.hstack([s4, s5, s6]),  hdiv(W3),
            np.hstack([s7, s8, s9]),  hdiv(W3),
            np.hstack([s10, s11, pad_p]),
        ])
        _, buf = cv2.imencode('.jpg', demo, [cv2.IMWRITE_JPEG_QUALITY, 85])
        yield (b'--frame\r\nContent-Type: image/jpeg\r\n\r\n' + buf.tobytes() + b'\r\n')

    cap.release()


# ============================================================
# ROUTES
# ============================================================

@app.route('/')
def index():
    return render_template('index.html', courses=COURSES)


# ── Verify ──────────────────────────────────────────────────
@app.route('/verify')
def verify():
    subject = request.args.get('subject', '')
    course  = request.args.get('course', '')
    # Validate subject belongs to course
    if course not in COURSES or subject not in COURSES.get(course, []):
        return redirect(url_for('index'))
    return render_template('verify.html', subject=subject, course=course)

@app.route('/verify_feed')
def verify_feed():
    subject = request.args.get('subject', 'General')
    course  = request.args.get('course',  'General')
    return Response(gen_frames_verify(subject, course),
                    mimetype='multipart/x-mixed-replace; boundary=frame')


# ── Admin login ─────────────────────────────────────────────
@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        if (request.form['email']    == ADMIN_EMAIL and
                request.form['password'] == ADMIN_PASSWORD):
            session['admin'] = True
            return redirect(url_for('register_page'))
        error = "Incorrect email or password."
    return render_template('login.html', error=error)

@app.route('/logout')
def logout():
    session.pop('admin', None)
    return redirect(url_for('index'))


# ── Register student ─────────────────────────────────────────
@app.route('/register')
def register_page():
    if not session.get('admin'):
        return redirect(url_for('login'))
    return render_template('register.html', courses=COURSES)

@app.route('/register_feed')
def register_feed():
    return Response(gen_frames_register(),
                    mimetype='multipart/x-mixed-replace; boundary=frame')

@app.route('/capture_face', methods=['POST'])
def capture_face():
    if not session.get('admin'):
        return redirect(url_for('login'))

    name    = request.form['student_name'].strip().replace(' ', '_')
    s_id    = request.form['student_id'].strip().replace('/', '-')
    course  = request.form['course']

    if course not in COURSES:
        return "<h3>Invalid course. <a href='/register'>Go Back</a></h3>"

    # Open fresh camera for capture
    import time
    cap = cv2.VideoCapture(0, cv2.CAP_DSHOW)
    if not cap.isOpened():
        cap = cv2.VideoCapture(0)
    for _ in range(5):   # flush buffer frames
        cap.read()
    ok, frame = cap.read()
    cap.release()

    if ok and frame is not None:
        # Filename: Name--StudentID--Course.jpg
        # Uses -- separator so underscores in student names don't break parsing
        course_fn = course.replace(' ', '-')
        filename  = f"{name}--{s_id}--{course_fn}.jpg"
        path      = os.path.join(IMAGE_FOLDER, filename)
        cv2.imwrite(path, frame)
        load_known_faces()
        flash(f"✅ Registered {name} for course: {course}", "success")
        return redirect(url_for('register_page'))

    flash("❌ Camera capture failed. Please try again.", "error")
    return redirect(url_for('register_page'))


# ── Dashboard ────────────────────────────────────────────────
@app.route('/dashboard')
def dashboard():
    """
    Shows attendance grouped by course → subject.
    For each subject shows a list of students who attended today
    (one row per student — no duplicates).
    Supports ?course=X and ?date=YYYY-MM-DD filters.
    """
    sel_course  = request.args.get('course', '')
    sel_subject = request.args.get('subject', '')
    sel_date    = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))

    try:
        date_obj = datetime.strptime(sel_date, '%Y-%m-%d')
    except ValueError:
        date_obj = datetime.now()

    day_start = date_obj.replace(hour=0,  minute=0,  second=0,  microsecond=0)
    day_end   = date_obj.replace(hour=23, minute=59, second=59, microsecond=999999)

    # Build query
    query = {"timestamp": {"$gte": day_start, "$lte": day_end}}
    if sel_course:
        query["course"] = sel_course
    if sel_subject:
        query["subject"] = sel_subject

    raw_logs = list(attendance_col.find(query).sort("timestamp", 1))

    # Deduplicate: keep only first occurrence per student+subject
    seen   = set()
    unique = []
    for log in raw_logs:
        key = (log['name'], log['subject'])
        if key not in seen:
            seen.add(key)
            unique.append(log)

    # Group: course → subject → [students]
    grouped = {}
    for course_name, subjects in COURSES.items():
        if sel_course and course_name != sel_course:
            continue
        grouped[course_name] = {}
        for subj in subjects:
            if sel_subject and subj != sel_subject:
                continue
            grouped[course_name][subj] = [
                r for r in unique
                if r.get('course') == course_name and r.get('subject') == subj
            ]

    total_present = len(unique)

    return render_template('dashboard.html',
                           grouped=grouped,
                           courses=COURSES,
                           sel_course=sel_course,
                           sel_subject=sel_subject,
                           sel_date=sel_date,
                           total_present=total_present)


# ── Preprocessing demo ───────────────────────────────────────
@app.route('/preprocessing')
def preprocessing():
    return render_template('preprocessing.html')

@app.route('/preprocessing_feed')
def preprocessing_feed():
    return Response(gen_preprocessing_demo(),
                    mimetype='multipart/x-mixed-replace; boundary=frame')


# ============================================================

# ============================================================
# EXCEL EXPORT
# /export_excel?date=YYYY-MM-DD&course=X&subject=Y
# Same filters as dashboard — downloads a .xlsx file
# ============================================================
@app.route('/export_excel')
def export_excel():
    sel_course  = request.args.get('course',  '')
    sel_subject = request.args.get('subject', '')
    sel_date    = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))

    try:
        date_obj = datetime.strptime(sel_date, '%Y-%m-%d')
    except ValueError:
        date_obj = datetime.now()

    day_start = date_obj.replace(hour=0,  minute=0,  second=0,  microsecond=0)
    day_end   = date_obj.replace(hour=23, minute=59, second=59, microsecond=999999)

    query = {"timestamp": {"$gte": day_start, "$lte": day_end}}
    if sel_course:  query["course"]  = sel_course
    if sel_subject: query["subject"] = sel_subject

    raw_logs = list(attendance_col.find(query).sort("timestamp", 1))

    # Deduplicate — same as dashboard
    seen, unique = set(), []
    for log in raw_logs:
        key = (log['name'], log['subject'])
        if key not in seen:
            seen.add(key)
            unique.append(log)

    # ── Build workbook ──────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # Styles
    thin_side   = Side(style="thin", color="AABBD4")
    cell_border = Border(left=thin_side, right=thin_side,
                         top=thin_side,  bottom=thin_side)
    center_al   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_al     = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    navy_fill  = PatternFill("solid", fgColor="0B1F45")
    blue_fill  = PatternFill("solid", fgColor="1565C0")
    light_fill = PatternFill("solid", fgColor="DCE8F8")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    grey_fill  = PatternFill("solid", fgColor="F4F7FB")

    # Row 1 — Main title
    ws.merge_cells("A1:F1")
    ws["A1"] = "FACE VERIFICATION ATTENDANCE SYSTEM"
    ws["A1"].font      = Font(name="Calibri", bold=True, size=15, color="FFFFFF")
    ws["A1"].fill      = navy_fill
    ws["A1"].alignment = center_al
    ws.row_dimensions[1].height = 26

    # Row 2 — Subtitle
    ws.merge_cells("A2:F2")
    ws["A2"] = "General Sir John Kotelawala Defence University"
    ws["A2"].font      = Font(name="Calibri", size=10, color="9EC5E8")
    ws["A2"].fill      = navy_fill
    ws["A2"].alignment = center_al
    ws.row_dimensions[2].height = 16

    # Row 3 — Filter info
    filter_text = f"Date: {sel_date}"
    if sel_course:  filter_text += f"   |   Course: {sel_course}"
    if sel_subject: filter_text += f"   |   Subject: {sel_subject}"
    filter_text += f"   |   Total Present: {len(unique)}"
    ws.merge_cells("A3:F3")
    ws["A3"] = filter_text
    ws["A3"].font      = Font(name="Calibri", size=10, bold=True, color="0B1F45")
    ws["A3"].fill      = light_fill
    ws["A3"].alignment = center_al
    ws.row_dimensions[3].height = 18

    # Row 4 — Column headers
    headers    = ["#",  "Student Name", "Course", "Subject", "Date",  "Time Marked"]
    col_widths = [5,     26,             22,        26,        14,      14          ]
    for i, (h, w) in enumerate(zip(headers, col_widths), start=1):
        c = ws.cell(row=4, column=i, value=h)
        c.font      = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        c.fill      = blue_fill
        c.alignment = center_al
        c.border    = cell_border
        ws.column_dimensions[c.column_letter].width = w
    ws.row_dimensions[4].height = 20

    # Data rows
    for idx, record in enumerate(unique, start=1):
        row = idx + 4
        ts  = record.get("timestamp")
        row_fill = white_fill if idx % 2 == 1 else grey_fill
        row_data = [
            idx,
            record.get("name",    ""),
            record.get("course",  ""),
            record.get("subject", ""),
            ts.strftime("%d %b %Y") if ts else "",
            ts.strftime("%H:%M:%S") if ts else "",
        ]
        for col, val in enumerate(row_data, start=1):
            c = ws.cell(row=row, column=col, value=val)
            c.font      = Font(name="Calibri", size=11)
            c.fill      = row_fill
            c.alignment = center_al if col in [1, 5, 6] else left_al
            c.border    = cell_border
        ws.row_dimensions[row].height = 17

    # Summary row
    sr = len(unique) + 6
    ws.merge_cells(f"A{sr}:E{sr}")
    ws.cell(row=sr, column=1, value=f"Total Students Present: {len(unique)}")
    ws.cell(row=sr, column=1).font      = Font(name="Calibri", bold=True, size=11, color="0B1F45")
    ws.cell(row=sr, column=1).fill      = light_fill
    ws.cell(row=sr, column=1).alignment = left_al
    ws.cell(row=sr, column=1).border    = cell_border
    ws.cell(row=sr, column=6,
            value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws.cell(row=sr, column=6).font      = Font(name="Calibri", size=10, italic=True, color="6B7A99")
    ws.cell(row=sr, column=6).fill      = light_fill
    ws.cell(row=sr, column=6).alignment = center_al
    ws.cell(row=sr, column=6).border    = cell_border
    ws.row_dimensions[sr].height = 18

    ws.freeze_panes = "A5"  # freeze header rows

    # Send as download
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    fname = f"Attendance_{sel_date}"
    if sel_course:  fname += f"_{sel_course.replace(' ','_')}"
    if sel_subject: fname += f"_{sel_subject.replace(' ','_')}"
    fname += ".xlsx"

    return send_file(output,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True,
                     download_name=fname)


if __name__ == '__main__':
    app.run(debug=True, port=5001, use_reloader=False)
